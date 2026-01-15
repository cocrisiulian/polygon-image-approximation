"""
Modul pentru exportul rezultatelor în format Excel.
Generează rapoarte comparative cu 3 foi de calcul:
- Foaia 1: Rezultate Generale
- Foaia 2: Detalii Simulated Annealing
- Foaia 3: Detalii Genetic Algorithm
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from datetime import datetime
from typing import Dict, List, Optional
import os


class ExcelExporter:
    """Exportă rezultatele algoritmilor în format Excel."""
    
    def __init__(self, output_path: str = "results/raport_comparative.xlsx"):
        """
        Inițializează exporterul Excel.
        
        Args:
            output_path: Calea către fișierul Excel de creat
        """
        self.output_path = output_path
        self.workbook = openpyxl.Workbook()
        
        # Eliminăm foaia implicită
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # Creăm cele 3 foi
        self.sheet_general = self.workbook.create_sheet("Rezultate Generale", 0)
        self.sheet_sa = self.workbook.create_sheet("Detalii SA", 1)
        self.sheet_ga = self.workbook.create_sheet("Detalii GA", 2)
        
        # Stiluri
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_results(self, 
                      image_name: str,
                      num_polygons: int,
                      sa_results: Optional[Dict] = None,
                      ga_results: Optional[Dict] = None):
        """
        Exportă rezultatele în Excel.
        
        Args:
            image_name: Numele imaginii procesate
            num_polygons: Numărul de poligoane
            sa_results: Dicționar cu rezultatele SA (history, fitness_final, time, etc.)
            ga_results: Dicționar cu rezultatele GA (history, fitness_final, time, etc.)
        """
        # Foaia 1: Rezultate Generale
        self._create_general_sheet(image_name, num_polygons, sa_results, ga_results)
        
        # Foaia 2: Detalii SA
        if sa_results:
            self._create_sa_sheet(sa_results)
        
        # Foaia 3: Detalii GA
        if ga_results:
            self._create_ga_sheet(ga_results)
        
        # Salvăm fișierul
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        self.workbook.save(self.output_path)
        print(f"\n📊 Raport Excel salvat: {self.output_path}")
    
    def _create_general_sheet(self, image_name: str, num_polygons: int,
                             sa_results: Optional[Dict], ga_results: Optional[Dict]):
        """Creează foaia cu rezultatele generale."""
        ws = self.sheet_general
        
        # Titlu
        ws['A1'] = '🎨 RAPORT COMPARATIV - APROXIMARE IMAGINI CU POLIGOANE'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Informații generale
        row = 3
        ws[f'A{row}'] = 'Data generării:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws[f'B{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Imagine procesată:'
        ws[f'B{row}'] = image_name
        ws[f'B{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Număr poligoane:'
        ws[f'B{row}'] = num_polygons
        ws[f'B{row}'].font = Font(bold=True)
        
        # Tabel comparativ
        row += 3
        ws[f'A{row}'] = 'TABEL COMPARATIV'
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        headers = ['Metrică', 'Simulated Annealing', 'Genetic Algorithm', 'Câștigător']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Date comparative
        metrics = []
        if sa_results and ga_results:
            sa_acc = 100 - sa_results.get('fitness_final', 0)
            ga_acc = 100 - ga_results.get('fitness_final', 0)
            
            metrics = [
                ('Acuratețe Finală (%)', f"{sa_acc:.2f}", f"{ga_acc:.2f}", 
                 'GA' if ga_acc > sa_acc else 'SA'),
                ('RMSE Final', f"{sa_results.get('fitness_final', 0):.2f}", 
                 f"{ga_results.get('fitness_final', 0):.2f}",
                 'GA' if ga_results.get('fitness_final', 0) < sa_results.get('fitness_final', 0) else 'SA'),
                ('Timp Total (secunde)', f"{sa_results.get('total_time', 0):.2f}", 
                 f"{ga_results.get('total_time', 0):.2f}",
                 'SA' if sa_results.get('total_time', 0) < ga_results.get('total_time', 0) else 'GA'),
                ('Iterații/Generații', str(sa_results.get('iterations', 0)), 
                 str(ga_results.get('generations', 0)), '-'),
            ]
        elif sa_results:
            sa_acc = 100 - sa_results.get('fitness_final', 0)
            metrics = [
                ('Acuratețe Finală (%)', f"{sa_acc:.2f}", '-', 'SA'),
                ('RMSE Final', f"{sa_results.get('fitness_final', 0):.2f}", '-', 'SA'),
                ('Timp Total (secunde)', f"{sa_results.get('total_time', 0):.2f}", '-', 'SA'),
                ('Iterații', str(sa_results.get('iterations', 0)), '-', 'SA'),
            ]
        elif ga_results:
            ga_acc = 100 - ga_results.get('fitness_final', 0)
            metrics = [
                ('Acuratețe Finală (%)', '-', f"{ga_acc:.2f}", 'GA'),
                ('RMSE Final', '-', f"{ga_results.get('fitness_final', 0):.2f}", 'GA'),
                ('Timp Total (secunde)', '-', f"{ga_results.get('total_time', 0):.2f}", 'GA'),
                ('Generații', '-', str(ga_results.get('generations', 0)), 'GA'),
            ]
        
        for metric_row in metrics:
            row += 1
            for col, value in enumerate(metric_row, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col == 4:  # Câștigător
                    cell.font = Font(bold=True, color="008000")
                    cell.alignment = Alignment(horizontal='center')
        
        # Ajustăm lățimea coloanelor
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
    
    def _create_sa_sheet(self, sa_results: Dict):
        """Creează foaia cu detalii Simulated Annealing."""
        ws = self.sheet_sa
        
        # Titlu
        ws['A1'] = '🔥 SIMULATED ANNEALING - DETALII'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Parametri
        row = 3
        ws[f'A{row}'] = 'PARAMETRI ALGORITM'
        ws[f'A{row}'].font = self.title_font
        
        row += 1
        params = [
            ('Temperatură inițială', sa_results.get('initial_temp', 100.0)),
            ('Rată răcire', sa_results.get('cooling_rate', 0.995)),
            ('Iterații maxime', sa_results.get('max_iterations', 10000)),
            ('Iterații executate', sa_results.get('iterations', 0)),
        ]
        
        for param_name, param_value in params:
            ws[f'A{row}'] = param_name
            ws[f'B{row}'] = param_value
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
        
        # Istoric iterații
        row += 2
        ws[f'A{row}'] = 'ISTORIC ITERAȚII'
        ws[f'A{row}'].font = self.title_font
        
        row += 1
        headers = ['Iterație', 'Fitness (RMSE)', 'Temperatură', 'Rate Acceptare (%)', 'Acuratețe (%)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Date istoric
        history = sa_results.get('history', [])
        for entry in history:
            row += 1
            iteration = entry.get('iteration', 0)
            fitness = entry.get('fitness', 0)
            temp = entry.get('temperature', 0)
            accept_rate = entry.get('acceptance_rate', 0) * 100
            accuracy = 100 - fitness
            
            data = [iteration, fitness, temp, accept_rate, accuracy]
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col > 1:  # Numerice
                    cell.number_format = '0.00'
        
        # Ajustăm lățimea coloanelor
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18
    
    def _create_ga_sheet(self, ga_results: Dict):
        """Creează foaia cu detalii Genetic Algorithm."""
        ws = self.sheet_ga
        
        # Titlu
        ws['A1'] = '🧬 GENETIC ALGORITHM - DETALII'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Parametri
        row = 3
        ws[f'A{row}'] = 'PARAMETRI ALGORITM'
        ws[f'A{row}'].font = self.title_font
        
        row += 1
        params = [
            ('Mărime populație', ga_results.get('population_size', 20)),
            ('Generații maxime', ga_results.get('max_generations', 1000)),
            ('Rată mutație', ga_results.get('mutation_rate', 0.1)),
            ('Rată crossover', ga_results.get('crossover_rate', 0.7)),
            ('Elitism', ga_results.get('elitism', 2)),
            ('Generații executate', ga_results.get('generations', 0)),
        ]
        
        for param_name, param_value in params:
            ws[f'A{row}'] = param_name
            ws[f'B{row}'] = param_value
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
        
        # Istoric generații
        row += 2
        ws[f'A{row}'] = 'ISTORIC GENERAȚII'
        ws[f'A{row}'].font = self.title_font
        
        row += 1
        headers = ['Generație', 'Best Fitness', 'Avg Fitness', 'Diversitate', 'Acuratețe (%)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Date istoric
        history = ga_results.get('history', [])
        for entry in history:
            row += 1
            generation = entry.get('generation', 0)
            best_fitness = entry.get('best_fitness', 0)
            avg_fitness = entry.get('avg_fitness', 0)
            diversity = entry.get('diversity', 0)
            accuracy = 100 - best_fitness
            
            data = [generation, best_fitness, avg_fitness, diversity, accuracy]
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col > 1:  # Numerice
                    cell.number_format = '0.00'
        
        # Ajustăm lățimea coloanelor
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18
